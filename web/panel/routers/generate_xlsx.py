from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from panel.models import MyModel
from datetime import datetime, timedelta
from django.utils import timezone
from io import BytesIO
from random import randint
from django.http import HttpResponse
from .base import BaseAPI, Get
from django.forms import Form
from django.http import JsonResponse
import json
import os
from ..forms import GetExel
from django.shortcuts import redirect


class Exel:
    def __init__(self, name_document:str):
        self.wb:Workbook = Workbook()
        self.wb.remove(self.wb.active)    
        self.name_document = name_document

    def generete_new_sheet(self, title):
        return self.wb.create_sheet(title=title)

    def make_file(self) -> str:
        self.set_alignment_for_workbook()
        path = f"web/static/exel_documents/{self.name_document}"
        self.wb.save(path)
        return path
    

    def set_alignment_for_workbook(self, horizontal='left', vertical='center', wrap_text=True):
        for sheet in self.wb.sheetnames:
            ws:Worksheet = self.wb[sheet]  
    
            for col_num, column in enumerate(ws.columns, 1):
                max_length = 0
                column = list(column)
                for cell in column:
                    if isinstance(cell.value, str):
                        for line in cell.value.splitlines():
                            max_length = max(max_length, len(line))
                    optimal_width = max_length + 5

                adjusted_width = max(optimal_width, ws.column_dimensions[get_column_letter(col_num)].width)
                ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.row == 1: 
                        cell.alignment = Alignment(horizontal="center", vertical=vertical, wrap_text=wrap_text)
                        


class GenerateExel(Exel):
    russian_weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']


    def create_document(self, start_time:datetime="",finish_time:datetime="", phrase:str = "", sites:list[str]=[]):
        news = MyModel.get_record_from_xlsx(start_time, finish_time, phrase, sites)
        if start_time and finish_time:
            self.generate_document_from_time_range(news, start_time, finish_time)
        else:
            self.generate_document(news)
        return self.make_file()

    def generate_document(self, news:list[MyModel]):
        sheet:Worksheet = self.generete_new_sheet(title="Новости")
        self.__fill_info(sheet, news)
    

    def generate_document_from_time_range(self, news_list:list[MyModel], start_time:datetime, finish_time:datetime):
        while start_time <= finish_time:
            russ_name_day = self.russian_weekdays[start_time.weekday()]
            formatted_date = start_time.strftime("%d.%m.%Y({})".format(russ_name_day))
            start_time += timedelta(days=1)
            sheet:Worksheet = self.generete_new_sheet(title=formatted_date)
            sorted_news = [news for news in news_list if news.time.date() == start_time.date()]
            self.__fill_info(sheet, sorted_news)


    def __fill_info(self, sheet:Worksheet, news_list:list[MyModel]):
        if not news_list:
            sheet.append(["Новостей нету"])
            return
        sheet.append(["Дата публікації", "Назва ресурсу", "Посилання", "Короткий зміст"])
        for news in news_list:
            sheet.append([news.time.strftime("%d-%m-%Y %H:%M:%S"), news.channel, news.chat_id, news.text.replace("None","") if news.text else "Текст по ссылке"])





class GetExelDocument(Get, BaseAPI):
    form_template: Form = GetExel


    def post_validate_data_range(self, start_time:str, finish_time:str):
        start_time = datetime.strptime(start_time, "%Y-%m-%d")
        finish_time = datetime.strptime(finish_time, "%Y-%m-%d")
        
        if start_time > finish_time:
            return "", ""
        return start_time,finish_time
    def handle_request(self) -> JsonResponse:

        current_time = datetime.now()

        start_time = self.form.data.get('dateFrom',"")
        finish_time = self.form.data.get('dateTo',"")
        phrase = self.form.data.get('keyword','')
        sites = self.form.data.get('category', '').split("; ") if self.form.data.get('category') else ''

        file_name = f"document_{current_time.strftime('%d-%m-%Y')}.xlsx"

        if start_time and finish_time:
            start_time,finish_time = self.post_validate_data_range(start_time,finish_time)



        exel = GenerateExel(name_document=file_name)
        exel.create_document(start_time=start_time,finish_time=finish_time , phrase=phrase, sites=sites)
        file_path = os.path.join('web', 'static', 'exel_documents', file_name)
        with open(file_path, 'rb') as f:
            file_data = f.read()
        response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
# class DelexelDocument()